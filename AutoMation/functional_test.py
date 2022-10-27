import time
import globals
import config
import logging
import telnet_powerswitch
import os
import datetime
import ShellHandler
import tkinter as tk
from tkinter import messagebox as mb

processurl = os.getcwd()
filename = input("Please enter filename:")
globals.str1 = filename
log_filename = datetime.datetime.now().strftime(filename + '_Main_%Y-%m-%d_%H_%M_%S.log')
if not os.path.exists('C:\\Python39\\Scripts\\Log'):
    os.makedirs('C:\\Python39\\Scripts\\Log')
os.mkdir('C:\\Python39\\Scripts\\Log\\'+log_filename, 777)
os.chdir('C:\\Python39\\Scripts\\Log\\'+log_filename)
logging.basicConfig(level = logging.DEBUG,
    format = '%(asctime)s %(name)-8s %(levelname)-8s %(message)s',
    datefmt = '%Y-%m-%d %H:%M:%S',
    filename = log_filename)
formatter = logging.Formatter('%(asctime)s %(name)-8s %(levelname)-8s %(message)s')
logger = logging.getLogger('Main:')
console = logging.StreamHandler()
console.setLevel(logging.INFO)
console.setFormatter(formatter)
logger.addHandler(console)

user_dir = ""

class Function_UserIO():

    def __init__(self):
        sam = ""
    def user_IXIAport(self):
        IXIAport = input("Please enter IXIA setting port.(ex:5678):")
        globals.str2 = IXIAport
    def user_ras_pi(self):
        ssh = ShellHandler.ShellHandler()
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        global user_dir
        user_dir = ssh.execute_some_command('ls /home')
        ssh.execute_some_command('exit')
        print ("user="+user_dir[2:-3])
        ### VERSION
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        version_log = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 0x1c 0x0 0x1e')
        ssh.execute_some_command('exit')
        version_start = version_log.index('value=')+6
        print ("version_log:"+str(version_log))
        print ("version_start:"+str(version_start))
        print (version_log[version_start:-3])
        f = open("version.txt", "a")
        f.write (version_log[version_start:-3])
        f.close()
        ### MII state
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        MII_log = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 0x1c 0x0 0x18')
        ssh.execute_some_command('exit')
        MII_start = version_log.index('value=')+8
        MII = MII_log[MII_start:-3]
        print (bin(int(MII, 16))[2:].zfill(16))
class Function_Report():
    def report(self, functionname, index, x1a_before, x1a_after):
        if not os.path.exists(functionname+".csv"):
            fp = open(functionname+".csv", "a")
            fp.write ("index,0x1a_before,0x1a_after"+"\n")
            fp.write (index+","+x1a_before+","+x1a_after+"\n")
            fp.close
        else:
            fp = open(functionname+".csv", "a")
            fp.write (index+","+x1a_before+","+x1a_after+"\n")
            fp.close
    def write_register(self, location, register_short_name, default_register, register_index, register_value):
        if not os.path.exists('default_register.xlsx'):
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "default_register"
            ws = wb['default_register']
            ws["A1"] = "register short name"
            ws['B1'] = "register"
            ws['C1'] = "default_register"
            ws['D1'] = "value"
            wb.save('default_register.xlsx')
        from openpyxl import load_workbook
        wb = load_workbook("default_register.xlsx")
        ws = wb.active
        ws.title = "default_register"
        ws = wb['default_register']
        ws['A'+location] = register_short_name
        ws['B'+location] = register_index
        ws['C'+localtion] = default_register
        ws['D'+location] = register_value
        wb.save('default_register.xlsx')  
class Functions():
    def __init__(self):
        globals.initialize()
        config.IXIA_CONFIG()
        config.POWERSWITCH_CONFIG()
        config.CABLESWITCH_CONFIG()
        config.nomal_variable()
        config.default_register()
        
    def selection_functions(self):
        print("Running case.\n\n Case 1 : LP mode test.\n Case 2 : WoL test.\n Case 3 : default value test.\n")
        print("If you want run case 1 case, please enter 1 , enter 0 means pass")
        Running_case = input("Please enter Running case:")
        print ("Running_case[0]="+Running_case[0])
        print ("Running_case[1]="+Running_case[1])
        if str(Running_case[0]) == "0" :
            logger.info ('LP : Ignore') 
        elif Running_case[0] == "1" :
            self.LP()
        if str(Running_case[1]) == "0" :
            logger.info ('WoL : Ignore') 
        elif Running_case[1] == "1" :
            self.WOL()
        if Running_case[2] == "0" :
            logger.info ('default register value : Ignore') 
        elif Running_case[2] == "1" :
            self.get_default_register_mdio()
    def power_cycle(self):
        telnet_client = telnet_powerswitch.TelnetClient()
        if telnet_client.login_host(config.host_ip,config.username,config.password):
            telnet_client.execute_some_command("sw o"+config.port+" off imme")
            time.sleep(5)  
            telnet_client.execute_some_command("sw o"+config.port+" on imme")
            telnet_client.logout_host()
    def LP(self):
        logger.info ('LP mode : Start')
        #### CablePlug case
        #### function.tcl
        #### option 0 : IXIA address
        #### option 1 : IXIA port1
        #### option 2 : IXIA port2
        #### option 3 : mode
        #### option 4 : masterslave
        #### option 5 : CSports
        #### option 6 : portspeed
        #### option 7 : numbers
        
        # step 1 : get register value(0x0 0x1)
        ssh = ShellHandler.ShellHandler()
        ssh.login_host(host = config.ArduinoIP2, user = "root", psw = "arduino")
        ssh.execute_some_command('./blink.py 8')
        ssh.execute_some_command('exit')
        time.sleep(3)
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        #/home/hpa/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 29 0x0 0x1 
        expect_value = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 1 0x0 0x1')
        #version_log = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 28 0x0 0x1e')
        ssh.execute_some_command('exit')
        expect_value_start = expect_value.index('value=')+6
        #print(expect_value)
        print ("before LP mode expect_value="+expect_value[expect_value_start:-3])
        
        # step 2 : set LP mode to 0xe
        
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        #/home/hpa/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 29 0x0 0x1 
        expect_value = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 1 0x0 0x14 0xe')
        #version_log = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 28 0x0 0x1e')
        ssh.execute_some_command('exit')
        expect_value_start = expect_value.index('value=')+6
        #print(expect_value)
        print ("0x14 value="+expect_value[expect_value_start:-11])
        
        
        
        
        # step 3 : plug out
        ssh.login_host(host = config.ArduinoIP2, user = "root", psw = "arduino")
        ssh.execute_some_command('./blink.py 7')
        ssh.execute_some_command('exit')
        time.sleep(15)
        
        # step 4 : get register value(0x0 0x1)
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        #/home/hpa/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 29 0x0 0x1 
        expect_value = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 1 0x0 0x1')
        #version_log = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 28 0x0 0x1e')
        ssh.execute_some_command('exit')
        expect_value_start = expect_value.index('value=')+6
        #print(expect_value)
        print ("after LP mode expect_value="+expect_value[expect_value_start:-3])
        
        # step 5 : plug in
        ssh.login_host(host = config.ArduinoIP2, user = "root", psw = "arduino")
        ssh.execute_some_command('./blink.py 8')
        ssh.execute_some_command('exit')
        time.sleep(5)
        
        # step 6 : get register value(0x0 0x1)
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        #/home/hpa/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 29 0x0 0x1 
        expect_value = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 1 0x0 0x1')
        #version_log = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 28 0x0 0x1e')
        ssh.execute_some_command('exit')
        expect_value_start = expect_value.index('value=')+6
        #print(expect_value)
        print ("over LP mode expect_value="+expect_value[expect_value_start:-3])
        
        # step 7 : set LP mode to 0x6
        # step 8: get register value(0x0 0x1)
        
        
        
        
        # for masterslave in config.MasterSlave :
        # #print (masterslave)
            # for csport in config.CSports :
                # print (csport)
                # ssh.login_host(host = config.ArduinoIP, user = "root", psw = "arduino")
                # ssh.execute_some_command('./blink.py ' + str(csport) )
                # ssh.execute_some_command('exit')
                # logger.info ('Ethernet switch1 port ' + str(csport) + ' on')
                # ssh.login_host(host = config.ArduinoIP2, user = "root", psw = "arduino")
                # ssh.execute_some_command('./blink.py ' + str(csport) )
                # ssh.execute_some_command('exit')
                # logger.info ('Ethernet switch2 port ' + str(csport) + ' on')
                # ssh.login_host(host = config.ArduinoIP3, user = "root", psw = "arduino")
                # ssh.execute_some_command('./blink.py ' + str(csport) )
                # ssh.execute_some_command('exit')
                # logger.info ('Ethernet switch3 port ' + str(csport) + ' on')
                # ssh.login_host(host = config.ArduinoIP4, user = "root", psw = "arduino")
                # ssh.execute_some_command('./blink.py ' + str(csport) )
                # ssh.execute_some_command('exit')
                # logger.info ('Ethernet switch4 port ' + str(csport) + ' on')
                # ### it is support 2 port cable swtich
                # for portspeed in config.Portspeeds :
                    # #print (portspeed)
                    # i = 1
                    # while i < config.testtimes :
                        # if portspeed == "speed100" and masterslave == "portSlave" :
                            # logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+ " igonre")
                            # print (masterslave+" "+str(csport)+" "+portspeed+" "+str(i))
                        # else :
                            # logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+ " start!")
                            # print (masterslave+" "+str(csport)+" "+portspeed+" "+str(i))
                            # #os.system("tclsh "+processurl+"\\function.tcl "+config.IXIA_IP+" "+globals.str2+" cableplug 1518 "+masterslave+" "+str(csport)+" "+portspeed+" "+str(i))
                            # os.system("tclsh "+processurl+"\\function.tcl "+config.IXIA_IP+" "+globals.str2+" cableplug 1518 5 1000 "+masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+" 150")
                            # logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i) + " end!")   
                        # i=i+1
        logger.info ('Case 1 : LP mode : End')
    def WOL(self):
    
        function_report = Function_Report()
        logger.info ('WoL : Start')
        logger.info ('WoL : case2.1.1 without password : start')    
        
        # self.power_cycle()
        
        # ssh = ShellHandler.ShellHandler()
        # ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        # ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe06 0x1')
                        
        # ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe08 0x9e35')
        # ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe09 0x56dc')
        # ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe0a 0x000d')
        
        # ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-write 99 0x1c 0x19 0x8000')
        # log_wol = ""
        # log_wol = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-read 99 0x1c 0x1a')
        # log_wol_start = log_wol.index('value=')+6
        # print ("log_wol_register="+log_wol[log_wol_start:-3])
        # register_before = log_wol[log_wol_start:-3]
        for masterslave in config.MasterSlave :
            for csport in config.CSports :
                for portspeed in config.Portspeeds :
                    i = 1
                    if portspeed == "speed100" and masterslave == "portSlave" :
                        logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+ " igonre")
                        print (masterslave+" "+str(csport)+" "+portspeed+" "+str(i))
                    else :
                        logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+ " start!")
                        print (masterslave+" "+str(csport)+" "+portspeed+" "+str(i))
                        # ### e. reset 0x1a value
                        os.system("tclsh "+processurl+"\\function.tcl "+config.IXIA_IP+" "+globals.str2+" wol 1500 5 3000 "+masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+" 150 0 0")
                        logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i) + " end!")   
        
        # step 3 confirm 0x1a value is 0x8000
        time.sleep(3)
        log_wol = ""
        log_wol = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-read 99 1 0x1a')
        log_wol_start = log_wol.index('value=')+6
        print ("log_wol_register="+log_wol[log_wol_start:-3])
        register_after = log_wol[log_wol_start:-3]
        ssh.execute_some_command('exit')
        
        function_report.report("WOL", "2.1.1", register_before, register_after)
        
        
        logger.info ('WoL : case2.1.1 without password : End')
        
        logger.info ('WoL : case2.1.2 with password : Start')
        
        self.power_cycle()               
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 1 0x1f 0xe06 0x5')
                        
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 1 0x1f 0xe08 0x9e35')
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 1 0x1f 0xe09 0x56dc')
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 1 0x1f 0xe0a 0x000d')
        
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 1 0x1f 0xe0b 0x0001')
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 1 0x1f 0xe0c 0x0203')
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 1 0x1f 0xe0d 0x0405')
        
        
        
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-write 99 1 0x19 0x8000')
        log_wol = ""
        log_wol = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-read 99 1 0x1a')
        log_wol_start = log_wol.index('value=')+6
        register_before = log_wol[log_wol_start:-3]
        print ("log_wol_register="+log_wol[log_wol_start:-3])
        
        for masterslave in config.MasterSlave :
            for csport in config.CSports :
                for portspeed in config.Portspeeds :
                    i = 1
                    if portspeed == "speed100" and masterslave == "portSlave" :
                        logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+ " igonre")
                        print (masterslave+" "+str(csport)+" "+portspeed+" "+str(i))
                    else :
                        logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+ " start!")
                        print (masterslave+" "+str(csport)+" "+portspeed+" "+str(i))
                        # ### e. reset 0x1a value
                        os.system("tclsh "+processurl+"\\function.tcl "+config.IXIA_IP+" "+globals.str2+" wol 1500 5 3000 "+masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+" 150 1 0")
                        logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i) + " end!")   
        
        # step 3 confirm 0x1a value is 0x8000
        time.sleep(3)
        log_wol = ""
        log_wol = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-read 99 0x1c 0x1a')
        log_wol_start = log_wol.index('value=')+6
        register_after = log_wol[log_wol_start:-3]
        print ("log_wol_register="+log_wol[log_wol_start:-3])
        
        ssh.execute_some_command('exit')
        function_report.report("WOL", "2.1.2", register_before, register_after)
        
        logger.info ('WoL : case2.1.2 with password : End')
        
        
        logger.info ('WoL : case2.2.1 IP format without password : Start')
        
        self.power_cycle()                
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe06 0x5')
                        
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe08 0x9e35')
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe09 0x56dc')
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe0a 0x000d')
        
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-write 99 0x1c 0x19 0x8000')
        log_wol = ""
        log_wol = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-read 99 0x1c 0x1a')
        log_wol_start = log_wol.index('value=')+6
        register_before = log_wol[log_wol_start:-3]
        print ("log_wol_register="+log_wol[log_wol_start:-3])
        
        for masterslave in config.MasterSlave :
            for csport in config.CSports :
                for portspeed in config.Portspeeds :
                    i = 1
                    if portspeed == "speed100" and masterslave == "portSlave" :
                        logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+ " igonre")
                        print (masterslave+" "+str(csport)+" "+portspeed+" "+str(i))
                    else :
                        logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+ " start!")
                        print (masterslave+" "+str(csport)+" "+portspeed+" "+str(i))
                        # ### e. reset 0x1a value
                        os.system("tclsh "+processurl+"\\function.tcl "+config.IXIA_IP+" "+globals.str2+" wol 1500 5 3000 "+masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+" 150 0 1")
                        logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i) + " end!")   
        
        # step 3 confirm 0x1a value is 0x8000
        time.sleep(3)
        log_wol = ""
        log_wol = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-read 99 0x1c 0x1a')
        log_wol_start = log_wol.index('value=')+6
        register_after = log_wol[log_wol_start:-3]
        print ("log_wol_register="+log_wol[log_wol_start:-3])
        
        ssh.execute_some_command('exit')
        function_report.report("WOL", "2.2.1", register_before, register_after)
        logger.info ('WoL : case2.2.1 IP format without password : End')
        
        logger.info ('WoL : case2.2.2 IP format with password : Start')
        
        self.power_cycle()               
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe06 0x5')
                        
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe08 0x9e35')
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe09 0x56dc')
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe0a 0x000d')
        
        
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe0b 0x0001')
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe0c 0x0203')
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-write 99 0x1c 0x1f 0xe0d 0x0405')
        
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-write 99 0x1c 0x19 0x8000')
        log_wol = ""
        log_wol = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-read 99 0x1c 0x1a')
        log_wol_start = log_wol.index('value=')+6
        register_before = log_wol[log_wol_start:-3]
        print ("log_wol_register="+log_wol[log_wol_start:-3])
        
        for masterslave in config.MasterSlave :
            for csport in config.CSports :
                for portspeed in config.Portspeeds :
                    i = 1
                    if portspeed == "speed100" and masterslave == "portSlave" :
                        logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+ " igonre")
                        print (masterslave+" "+str(csport)+" "+portspeed+" "+str(i))
                    else :
                        logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+ " start!")
                        print (masterslave+" "+str(csport)+" "+portspeed+" "+str(i))
                        # ### e. reset 0x1a value
                        os.system("tclsh "+processurl+"\\function.tcl "+config.IXIA_IP+" "+globals.str2+" wol 1500 5 3000 "+masterslave+" "+str(csport)+" "+portspeed+" "+str(i)+" 150 1 1")
                        logger.info (masterslave+" "+str(csport)+" "+portspeed+" "+str(i) + " end!")   
        
        # step 3 confirm 0x1a value is 0x8000
        time.sleep(3)
        log_wol = ""
        log_wol = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-read 99 0x1c 0x1a')
        log_wol_start = log_wol.index('value=')+6
        register_after = log_wol[log_wol_start:-3]
        print ("log_wol_register="+log_wol[log_wol_start:-3])
        ssh.execute_some_command('exit')
        function_report.report("WOL", "2.2.2", register_before, register_after)
        logger.info ('WoL : case2.2.2 IP format with password : End')
        logger.info ('Case 2 : WoL : End')
    def get_default_register_mdio(self):
        function_report = Function_Report()
        logger.info ('WoL : Start')
        logger.info ('WoL : case3 default register: start')    
        # self.power_cycle()
        ### idle register value
        idx = 2
        ssh = ShellHandler.ShellHandler()
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        for register, register_short_name, default_register in zip(config.register, config.register_short_name, config.default_register):
            print ("register len="+str(len(register)))
            i = 1
            while i < len(register):
                time.sleep(1) 
                log_reg = ""
                log_reg = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 0x1c '+register[0]+" "+register[i])
                print ("log_reg==="+log_reg)
                log_reg_start = log_reg.index('value=')+6
                
                print ("log_reg_register="+log_reg[log_reg_start:-3])
                register_id = register[0]+" "+register[i]
                #print (register_id)
                function_report.write_register(str(idx), register_short_name[i], default_register[i], register_id, log_reg[log_reg_start:-3])
                i=i+1
                idx = idx+1
        ssh.execute_some_command('exit')
        
        ### hw reset register value
        idx = 122
        ssh = ShellHandler.ShellHandler()
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        ### hw reset step
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/gpio22_rst')
        time.sleep(5) 
        
        for register, register_short_name, default_register in zip(config.register, config.register_short_name, config.default_register):
            print ("register len="+str(len(register)))
            i = 1
            while i < len(register):
                time.sleep(1) 
                log_reg = ""
                log_reg = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 0x1c '+register[0]+" "+register[i])
                print ("log_reg==="+log_reg)
                log_reg_start = log_reg.index('value=')+6
                
                print ("log_reg_register="+log_reg[log_reg_start:-3])
                register_id = register[0]+" "+register[i]
                #print (register_id)
                function_report.write_register(str(idx), register_short_name[i], default_register[i], register_id, log_reg[log_reg_start:-3])
                i=i+1
                idx = idx+1
        ssh.execute_some_command('exit')
        
        ### sw reset register value
        idx = 242
        ssh = ShellHandler.ShellHandler()
        ssh.login_host(host = config.RasPiIP, user = "root", psw = "12345678")
        ### hw reset step
        ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c22-write 99 0x1c 0x0 0xb040')
        time.sleep(5) 
        
        for register, register_short_name, default_register in zip(config.register, config.register_short_name, config.default_register):
            print ("register len="+str(len(register)))
            i = 1
            while i < len(register):
                time.sleep(1) 
                log_reg = ""
                log_reg = ssh.execute_some_command('/home/'+user_dir[2:-3]+'/Desktop/ras_pi/mypg_ethswbox/ethswbox/build/fapi-mdio-c45-read 99 0x1c '+register[0]+" "+register[i])
                print ("log_reg==="+log_reg)
                log_reg_start = log_reg.index('value=')+6
                
                print ("log_reg_register="+log_reg[log_reg_start:-3])
                register_id = register[0]+" "+register[i]
                #print (register_id)
                function_report.write_register(str(idx), register_short_name[i], default_register[i], register_id, log_reg[log_reg_start:-3])
                i=i+1
                idx = idx+1
        ssh.execute_some_command('exit')
        
        
if __name__ == "__main__":
    functions = Functions()
    function_userio = Function_UserIO()
    function_userio.user_IXIAport()
    function_userio.user_ras_pi()
    functions.selection_functions()